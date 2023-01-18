import requests
import time
import json
import string
import concurrent.futures
from tqdm import tqdm
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter


class HX_missingHosts(Exception):
    pass

def make_request(method, endpoint, headers, params, proxies, timeout, verify): #
    """
     Description
    ----------
    makes a GET or POST request. Retries 5 times if unsuccessful 
    
     Parameters
    ----------
    query : str
    
     Returns
    -------
    'None' if it is unable to make a successful Helix search, otherwise a list of items in the groupby search
    
     Examples
    --------

    """
    retries = 5
    num_of_tries = 0
    while num_of_tries < retries:
        try:
            session = requests.Session()
            if(method == "GET"):
                response = session.get(endpoint, headers=headers, params=params, verify=verify, timeout=timeout, proxies=proxies)
            elif(method == "POST"):
                response = session.post(endpoint, headers=headers, data=params, verify=verify, timeout=timeout, proxies=proxies)
            else:
                #sys.stderr.write("make_request method ERROR. Unsupported method: {}".format(method))
                return None
            response.raise_for_status()
            return response
        except Exception as e:
            err = str(e)
            try:
                if hasattr(e, "response"):
                    try:
                        response = e.response.json()
                        err = "API Error:"
                        detail_error = response.get("detail", None)
                        if detail_error is not None:
                            err += " {}".format(str(detail_error))
                        error = response.get("error", None)
                        if error is not None:
                            err += " {}".format(str(error))
                        message_error = response.get("message", None)
                        if message_error is not None:
                            err += " {}".format(str(message_error))
                    except ValueError:
                        err = str(e)
            except AttributeError:
                err = str(e)
            #sys.stderr.write(str(err))
            print(err)
            num_of_tries += 1
            time.sleep(1)
            #exit(-1)




def collect_all_entries(apikey, api_server, endpoint, debug=False):
    if debug == True:
        start_time = datetime.utcnow()

    headers = {
        "x-fireeye-api-key": apikey,
        "Accept": "application/json",
        "Content-Type": "application/json"
    }
    
    entries = []
    #collect first 100 entries - starting at offset 0
    #default offset is 0 - default limit is 50
    raw_json = make_request(method="GET", endpoint="{}{}?offset=0&limit=50".format(api_server, endpoint), headers=headers, params="", proxies="", timeout=180, verify=True).json()

    if raw_json["data"]["total"] == 0:
        return []
    
    for entry in raw_json["data"]["entries"]:
        entries.append(entry.copy())
    
    offset = 50
    total = raw_json["data"]["total"]

    while raw_json["data"]["entries"] != []:
        raw_json = make_request(
                method="GET", 
                endpoint="{}{}?offset={}".format(api_server, endpoint, offset), 
                headers=headers, 
                params="", 
                proxies="", 
                timeout=180, 
                verify=True).json()
        for entry in raw_json["data"]["entries"]:
            entries.append(entry.copy())
        offset += 50
        #print(len(entries))
    if debug == True:
        end_time = datetime.utcnow()
        delta = end_time - start_time
        print("Endpoint: {}".format(endpoint))
        print("Total: {}".format(total))
        print("Total Collected: {}".format(len(entries)))
        print("Start Time: {}; End Time: {}".format(start_time, end_time))
        print("Total Time: {}".format(delta))
        print("################################################")
    
    return entries

def main(apikey, force_data_collection):
    start_time = datetime.utcnow()
    
    #####REQUIRED VARIABLES#####
    headers = {
        "x-fireeye-api-key": apikey,
        "Accept": "application/json",
        "Content-Type": "application/json"
    }
    
    api_server = "https://hexABC123-hx-webui-1.hex07.helix.apps.fireeye.com/hx/api/v3"
    
    ############################
    
    


    #code checks if the hosts.json file exists and is a valid json file. 
    #Sample Data:
    """
    {
        "api_server": "123",
        "hosts": [
            {
                "_id": "123",
                [..........]
            },
            {
                "_id": "124",
                [..........]
            }
        ]
    }
    """
    if force_data_collection == False:
        file_name = "hosts.json"
        hosts_json_exception_met = False
        print("[{}] Importing {}...".format(datetime.utcnow(), file_name))


        try:
            with open(file_name, "r") as file:
                data = json.load(file)
        
        except FileNotFoundError as file_err:
            print("[{}] hosts.json not found".format(datetime.utcnow()))
            hosts_json_exception_met = True
            pass
        
        except json.decoder.JSONDecodeError as json_err:
            print("[{}] ERROR: {} is not a valid JSON file.".format(datetime.utcnow(), file_name))
            print("[{}] ERROR: {}".format(datetime.utcnow(), json_err))
            print("[{}] ERROR: The script will need to re-collect the data from HX. This may take some time.".format(datetime.utcnow()))
            def should_i_stay_or_should_i_go():
                choice = input("Would You Like to Start Re-collecting the Data? [y/n]: ").lower()
                yes = {'yes','y', 'ye'}
                no = {'no','n', 'nay'}

                if choice in yes:
                    hosts_json_exception_met == True
                    return
                elif choice in no:
                    print("Okay Bye Bye")
                    exit()
                else:
                    print("Please respond with 'yes' or 'no'")
                    should_i_stay_or_should_i_go()

            should_i_stay_or_should_i_go()
            pass
    else:
        hosts_json_exception_met == True

    if (hosts_json_exception_met == True) or (data["api_server"] != api_server):
        print("[{}] Starting Data Collection".format(datetime.utcnow()))
        print("[{}] Collecting all hosts from HX...".format(datetime.utcnow()))
        raw_hosts = collect_all_entries(apikey, api_server, endpoint="/hosts")
    
        print("[{}] Collecting each hosts configuration from HX...".format(datetime.utcnow()))
        #a = datetime.utcnow()
        hosts = []
    
        def get_host_config(host):
            config = make_request(
                method="GET",
                endpoint="{}/hosts/{agent_id}/configuration/actual.json".format(api_server, agent_id=host["_id"]),
                headers=headers,
                params="", 
                proxies="", 
                timeout=180, 
                verify=True).json()
            return config

        
        #this runs the 'get_host_config' function for each host in raw_hosts, and adds the results to the host.
        #for every host in raw_hosts, the 'get_host_config' function will be executed as an independant thread
        with concurrent.futures.ThreadPoolExecutor() as executor: 
            futures_to_items = {executor.submit(get_host_config, item): item for item in raw_hosts}
            for future in concurrent.futures.as_completed(futures_to_items):
                item = futures_to_items[future]
                try:
                    result = future.result()
                    item["config"] = result
                    hosts.append(item)
                except Exception as exc:
                    # Handle any exceptions here
                    print(exc)
                    pass
        
        with open('hosts.json', 'w') as f:
            initialize = {
                "api_server": api_server,
                "hosts": hosts
            } 
            f.write(json.dumps(initialize, indent=4))
    else:
        hosts = data["hosts"]
    


    
    print("[{}] Exporting data...".format(datetime.utcnow()))

    #Export Data to Excel File
    #https://openpyxl.readthedocs.io/en/stable/

    wb = Workbook()
    dest_filename = 'hx-map.xlsx'

    print("[{}] Exporting Worksheet1: 'Hosts'...".format(datetime.utcnow()))
    worksheet1 = wb.active
    worksheet1.title = "All Hosts"
    worksheet1["A1"] = "Hostname"
    worksheet1["B1"] = "IP"
    worksheet1["C1"] = "Operating System"
    worksheet1["D1"] = "Agent Version"
    worksheet1["E1"] = "Last Check-in Time"
    worksheet1["F1"] = "Real-Time Enabled"
    worksheet1["G1"] = "Real-Time File or Folder Exclusions"
    worksheet1["H1"] = "Real-Time Process Exclusions"
    worksheet1["I1"] = "Exploit Guard Enabled"
    worksheet1["J1"] = "Exploit Guard File Exclusions"
    worksheet1["K1"] = "Exploit Guard Folder Exclusions"
    worksheet1["L1"] = "Exploit Guard MD5 Exclusions"
    worksheet1["M1"] = "Malware Signature & Heuristics Enabled"
    worksheet1["N1"] = "Malware Signature & Heuristic Quarantine Enabled"
    worksheet1["O1"] = "Malware Guard Enabled"
    worksheet1["P1"] = "Malware Guard Quarantine Enabled"
    worksheet1["Q1"] = "Malware Detection File or Folder Exclusions"
    worksheet1["R1"] = "Malware Detection Process Exclusions"
    worksheet1["S1"] = "Malware Detection MD5 Exclusions"
    #worksheet1["T1"] = "EventStreamer Enabled" #Possibly to be added - probably never

    starting_row_number = 2 #used to account for the first row being used for Column Names. Rows in Excel start with 1, not 0
    
    
    def exclusion_list_export(column, exclusion_list):
        row = 0
        for file_exclusion in exclusion_list:
            row = row + starting_row_number
            worksheet1["{}{}".format(column,row)] = file_exclusion
            row += 1
    
    def determine_next_starting_row():
        #Determine next starting row number
            #calculate which column uses the most rows, make the next starting row number the most rows + 1
        """
        THIS SECTION WILL ONLY WORK IF THE COLUMNS DO NOT EXCEED PAST "Z"
        column_prefixes is a generated list of uppercase letters from A-Z
        """
        column_prefixes = list(string.ascii_uppercase)
        largest_row = 0
        for column in column_prefixes:
            x = True
            row_num = starting_row_number
            while x:

                row_to_check = worksheet1["{}{}".format(column, row_num)]
                #print("Column {}, Row {} == {}".format(column, row_num, row_to_check.value))
                
                if row_to_check.value is not None:
                    #print("{} is not None. Going to Next Row".format(row_to_check.value))
                    row_num += 1
                
                elif row_to_check.value is None:
                    #print("{} is None. Stopping While Loop".format(row_to_check.value))
                    x = False
                    if row_num > largest_row:
                        largest_row = row_num
        return largest_row
    
    def latest_timestamp(timestamp1, timestamp2):
        datetime1 = datetime.strptime(timestamp1, "%Y-%m-%dT%H:%M:%S.%fZ")
        datetime2 = datetime.strptime(timestamp2, "%Y-%m-%dT%H:%M:%S.%fZ")
        if datetime1 > datetime2:
            return timestamp1
        else:
            return timestamp2
    
    for host in hosts:
        #check if all required keys are present
        #check_keys_exists(
        #    json_obj=json.dumps(host), 
        #    keys=[
        #        "config", 
        #        "hostname", 
        #        "primary_ip_address", 
        #        "os", 
        #        "events", 
        #        "malwareDetection", 
        #        "exploitDetection"
        #    ]
        #)
        column_prefixes = list(string.ascii_uppercase)
        for column in column_prefixes:
            
            #change the Header colour and embolden the font
            worksheet1["{}{}".format(column, 1)].fill = PatternFill(start_color='949699', end_color='949699', fill_type="solid")
            worksheet1["{}{}".format(column, 1)].font = Font(bold=True)

            #change the starting row to a colour, which will easily identify when a new host starts
            worksheet1["{}{}".format(column, starting_row_number)].fill = PatternFill(start_color='4287f5', end_color='4287f5', fill_type="solid")
        
        #input basic data for host
        worksheet1["A{}".format(starting_row_number)] = host["hostname"]
        worksheet1["B{}".format(starting_row_number)] = host["primary_ip_address"]
        worksheet1["C{}".format(starting_row_number)] = host["os"]["product_name"]
        worksheet1["D{}".format(starting_row_number)] = host["agent_version"]
        worksheet1["E{}".format(starting_row_number)] = latest_timestamp(
            timestamp1=str(host["last_audit_timestamp"]), 
            timestamp2=str(host["last_poll_timestamp"])
        )


        
        def exclusion_list_export(column, exclusion_list):
            if exclusion_list != []:
                row = starting_row_number
                for file_exclusion in exclusion_list:
                    worksheet1["{}{}".format(column,row)] = file_exclusion
                    row += 1
        
        
        """
        #write each policy name applied to the host on a new row 
        row = 0 #create a new row variable to ensure starting_row_number stays the same
        for policy in host["policies"]:
            row = row + starting_row_number
            worksheet1["D{}".format(row)] = policy["name"]
            row += 1

        #write each host-set name applied to the host on a new row 
        row = 0
        for set in host["host-sets"]:
            row = row + starting_row_number
            worksheet1["E{}".format(row)] = set["name"]
            row += 1
        """

        #REAL-TIME
        """
        worksheet1["F1"] = "Real-Time Enabled"
        worksheet1["G1"] = "Real-Time File or Folder Exclusions"
        worksheet1["H1"] = "Real-Time Process Exclusions"
        """
        if "events" in host["config"]:
            realTime = host["config"]["events"]

        if realTime["active_collection_enabled"] == True:
            worksheet1["F{}".format(starting_row_number)] = "True"
            exclusion_list_export(column="G", exclusion_list=realTime["excludedPaths"])
            exclusion_list_export(column="H", exclusion_list=realTime["excludedProcessNames"])
        else:
            worksheet1["F{}".format(starting_row_number)] = "False"


        #Exploit Guard
        """
        worksheet1["I1"] = "Exploit Guard Enabled"
        worksheet1["J1"] = "Exploit Guard File Exclusions"
        worksheet1["K1"] = "Exploit Guard Folder Exclusions"
        worksheet1["L1"] = "Exploit Guard MD5 Exclusions"
        """
        exploitGuard = host["config"]["exploitDetection"]

        def dump_exploit_guard_exclusions():
            if "excludedFiles" in exploitGuard:
                exclusion_list_export(column="J", exclusion_list = exploitGuard["excludedFiles"])
            if "excludedPaths" in exploitGuard:
                exclusion_list_export(column="K", exclusion_list = exploitGuard["excludedPaths"])
            if "excludedMD5s" in exploitGuard:
                exclusion_list_export(column="L", exclusion_list = exploitGuard["excludedMD5s"])

        if host["os"]["platform"] == "win": #exploit guard is only available for windows
            if exploitGuard["enable_protection"] == True:
                if exploitGuard["enable_production"] == True:
                    if ("Server" in host["os"]["product_name"]) or ("server" in host["os"]["product_name"]): #Exploit Guard can be disabled specifically for Windows Servers
                        if exploitGuard["enable_server_os"] == True:
                            worksheet1["I{}".format(starting_row_number)] = "True"
                            dump_exploit_guard_exclusions()
                    else:
                        worksheet1["I{}".format(starting_row_number)] = "True" #all windows workstations will have exploit guard enabled when the 'enable_production' switch is set to True
                        dump_exploit_guard_exclusions()
            if worksheet1["I{}".format(starting_row_number)].value != True:
                worksheet1["I{}".format(starting_row_number)] = "False"
        else:
            worksheet1["I{}".format(starting_row_number)] = "Not Supported for {}".format(host["os"]["platform"])
            worksheet1["J{}".format(starting_row_number)] = "Not Supported for {}".format(host["os"]["platform"])
            worksheet1["K{}".format(starting_row_number)] = "Not Supported for {}".format(host["os"]["platform"])
            worksheet1["L{}".format(starting_row_number)] = "Not Supported for {}".format(host["os"]["platform"])




        #Malware Detection
        """
        worksheet1["M1"] = "Malware Signature & Heuristics Enabled"
        worksheet1["N1"] = "Malware Signature & Heuristic Quarantine Enabled"
        worksheet1["O1"] = "Malware Guard Enabled"
        worksheet1["P1"] = "Malware Guard Quarantine Enabled"
        worksheet1["Q1"] = "Malware Detection File or Folder Exclusions"
        worksheet1["R1"] = "Malware Detection Process Exclusions"
        worksheet1["S1"] = "Malware Detection MD5 Exclusions"
        """

        malwareDetection = host["config"]["malwareDetection"]

        def dump_malware_detection_exclusions():
            try:
                exclusion_list_export(column="Q", exclusion_list = malwareDetection["excludedFiles"])
                exclusion_list_export(column="S", exclusion_list = malwareDetection["excludedMD5s"])
                exclusion_list_export(column="R", exclusion_list = malwareDetection["excludedProcesses"])
            except KeyError:
                None
        
        #Signature and Heuristic Detection
        if malwareDetection["enable"] == True: #determines if Signature and Heuristic is enabled
            worksheet1["M{}".format(starting_row_number)] = "True"
            if malwareDetection["quarantine"]["enable"] == True: #determine if Signature and Heuristic quarantine is enabled
                worksheet1["N{}".format(starting_row_number)] = "True"
            else:
                worksheet1["N{}".format(starting_row_number)] = "False"

        else:
            worksheet1["M{}".format(starting_row_number)] = "False"
            worksheet1["N{}".format(starting_row_number)] = "False"
        
        #Malware Guard Detection
        #WINDOWS
        if host["os"]["platform"] == "win": 

            if malwareDetection["engine_configuration"]["mg"]["enable"] == True: #determines if Malware Guard is enabled
                worksheet1["O{}".format(starting_row_number)] = "True"
                if malwareDetection["engine_configuration"]["mg"]["quarantine_enable"] == True: #determine if Malware Guard quarantine is enabled
                    worksheet1["P{}".format(starting_row_number)] = "True"
                else:
                    worksheet1["P{}".format(starting_row_number)] = "False"
            else:
                worksheet1["O{}".format(starting_row_number)] = "False"
                worksheet1["P{}".format(starting_row_number)] = "False"

        else: #Malware Guard is only available for Windows
            worksheet1["O{}".format(starting_row_number)] = "Not Supported for {}".format(host["os"]["platform"]) 
            worksheet1["P{}".format(starting_row_number)] = "Not Supported for {}".format(host["os"]["platform"])
            worksheet1["R{}".format(starting_row_number)] = "Not Supported for {}".format(host["os"]["platform"])


        starting_row_number = determine_next_starting_row() #this will tell the script where to start dumping data again, should a host need multple rows to dump all its exlcusions

        
    print("[{}] Saving Excel File...".format(datetime.utcnow()))
    wb.save(filename = dest_filename)
    
    end_time = datetime.utcnow()
    print("[{}] Total Execution Time: {}".format(end_time, end_time - start_time))
        
    #TODO
    # add count of all hosts
    # add progress bar for when the script needs to collect all hosts & host configurations
    # use threading for collecting all hosts
        # can allocate the total amount of hosts equally across the maximum amount of threads
            #Examlple:
            #   100 threads max 
            #   1000 hosts to collect
            #   each thread will handle collecting 10 hosts
                #   Thread 1 will collect hosts 0 - 9
                #   Thread 2 will collect hosts 10 - 19
                #   Thread 3 will collect hosts 20 - 29
                #   [....]
                #   Thread 10 will collect hosts 90 - 99
    # add column for each policy applied to host
    # add column for each host-set applied to host
if __name__ == "__main__":
    main(
        apikey = "",
        force_data_collection=False #set this to True if you want the script to always collect fresh data
    )
